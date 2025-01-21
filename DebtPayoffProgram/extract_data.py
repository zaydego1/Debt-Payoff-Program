from asyncio.windows_events import NULL
import pandas as pd
from openpyxl import load_workbook

"""Class used for data extraction"""
class ExtractData:

    def __init__(self, spreadsheet):
        self.spreadsheet = str(spreadsheet) + ".xlsx"
        self.wb = load_workbook(self.spreadsheet, read_only=True)
  
    """Validate that Credit Card Data table exists and return it if so."""
    def extractCreditCard(self):

        if not 'Credit Card Data' in self.wb.sheetnames:
            raise Exception("Sheet: 'Credit Card Data' not found in Excel file. Please add sheet " + 
                "or rename sheet correctly.")
        credit_card_df = pd.read_excel(self.spreadsheet, 'Credit Card Data')       
        return credit_card_df

    """Validate that Loan Data table exists and return it if so."""
    def extractLoanData(self):

        if not 'Loan Data' in self.wb.sheetnames:
            raise Exception("Sheet: 'Loan Data' not found in Excel file. Please add sheet " + 
                "or rename sheet correctly.")
        loan_df = pd.read_excel(self.spreadsheet, 'Loan Data')       
        return loan_df
        

    